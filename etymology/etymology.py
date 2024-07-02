import pandas as pd
import yaml
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
import os

def classify_vehicle_type(input_path, sheet_names, columns_to_check_map, output_path, new_column_name):
    writer = pd.ExcelWriter(output_path, engine='openpyxl')

    # 读取 Excel 文件以获取所有 sheet 名称
    xls = pd.ExcelFile(input_path)

    for sheet_name in sheet_names:
        if sheet_name not in xls.sheet_names:
            print(f"Sheet '{sheet_name}' 不存在，跳过该 sheet")
            continue

        df = pd.read_excel(input_path, sheet_name=sheet_name)
        print(f"Processing sheet: {sheet_name}")
        print(f"DataFrame columns: {df.columns.tolist()}")

        # 检查新列名是否存在，如果不存在则添加
        if new_column_name not in df.columns:
            df[new_column_name] = None

        print(f"New column name to be used: {new_column_name}")

        # 遍历指定的列，检查是否包含关键字
        for index, row in df.iterrows():
            if pd.notna(df.at[index, new_column_name]):
                continue  # 如果新列已有值，跳过此行

            for column, keyword_map in columns_to_check_map.items():
                if column not in df.columns:
                    print(f"Column '{column}' 不存在于 sheet '{sheet_name}'，跳过该列")
                    continue

                cell_value = str(row[column])
                matched = False
                for keyword, vehicle_type in keyword_map.items():
                    if keyword in cell_value:
                        df.at[index, new_column_name] = vehicle_type
                        matched = True
                        break
                if matched:
                    break

            if pd.isna(df.at[index, new_column_name]):
                df.at[index, new_column_name] = '其他'

        # 将 DataFrame 写入 Excel 文件
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    writer.book.save(output_path)

    # 打开写入业务类型列后的 Excel 文件，添加数据验证
    wb = load_workbook(output_path)

    for sheet_name in sheet_names:
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]

        # 获取所有业务类型用于数据验证
        all_types = list(set([item for sublist in columns_to_check_map.values() for item in sublist.values()])) + ['其他']
        dv = DataValidation(type="list", formula1=f'"{",".join(all_types)}"', showDropDown=True)

        # 获取 '业务类型' 列的列号
        type_col_idx = len(df.columns)  # 新列在最后，获取最后一列的索引

        # 添加数据验证到业务类型列
        for row in range(2, ws.max_row + 1):
            cell = f"{chr(64 + type_col_idx)}{row}"  # chr(65) is 'A'
            dv.add(ws[cell])

        ws.add_data_validation(dv)

    wb.save(output_path)


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description='对 Excel 文件的某几列进行关键字匹配，并添加业务类型列')
    parser.add_argument('-c', '--config_file', help='配置文件的路径', default='config.yaml')

    args = parser.parse_args()

    config_file = args.config_file

    # 检查配置文件是否存在
    if not os.path.exists(config_file):
        raise FileNotFoundError(f"配置文件 {config_file} 不存在")

    # 读取配置文件
    with open(config_file, 'r', encoding='utf-8') as file:
        config = yaml.safe_load(file)

    input_path = config['input_path']
    sheet_names = config['sheet_names']
    columns_to_check_map = config['columns_to_check_map']
    output_path = config['output_path']
    new_column_name = config['new_column_name']

    classify_vehicle_type(input_path, sheet_names, columns_to_check_map, output_path, new_column_name)

    print("Success!")
    input("Press Enter to exit...")
